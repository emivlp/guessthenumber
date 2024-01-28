import secrets
import getpass

import openpyxl
from openpyxl import Workbook

import matplotlib.pyplot as plt


# Función para generar un número aleatorio
def num_guess():
    return secrets.randbelow(1000) + 1


# Nivel de Dificultad

def obtener_dificultad():
    while True:
        print("Seleccione el nivel de dificultad:")
        print()
        print("1. Fácil (20 intentos)")
        print("2. Medio (12 intentos)")
        print("3. Difícil (5 intentos)")
        try:
            opcion_dificultad = int(input("El nivel de dificultad a elegir es: "))
            if opcion_dificultad in [1, 2, 3]:
                return opcion_dificultad
            else:
                print("Opción no válida. Por favor, elija una opción del 1 al 3.")
                print()
        except ValueError:
            print("Entrada no válida. Ingrese un número.")

# Estadísticas excel

# Diccionario para almacenar las estadísticas de cada jugador
estadisticas_jugadores = {}


def guardar_estadisticas(nombre_jugador, resultado):
    try:
        excel_guess_statistics = openpyxl.load_workbook("/Users/emvlp/Desktop/guess_the_number_statistics.xlsx")
    except FileNotFoundError: # Si el archivo no exite lo crea openpyxl
        excel_guess_statistics = Workbook()

    sheet_guess1 = excel_guess_statistics.active

    # Si el archivo está vacío, añade encabezados
    if sheet_guess1["A1"].value is None:
        sheet_guess1.append(["Nombre del Jugador", "Resultado de la partida"])

    sheet_guess1.append([nombre_jugador, resultado])

    excel_guess_statistics.save("/Users/emvlp/Desktop/guess_the_number_statistics.xlsx")



    # Se actualiza el diccionario de estadísticas de los jugadores
    if nombre_jugador not in estadisticas_jugadores:
        estadisticas_jugadores[nombre_jugador] = {"Ganadas": 0, "Perdidas": 0}

    if resultado == "Partida Ganada":
        estadisticas_jugadores[nombre_jugador]["Ganadas"] += 1
    elif resultado == "Partida Perdida":
        estadisticas_jugadores[nombre_jugador]["Perdidas"] += 1
    

def mostrar_estadisticas():
    try:
        excel_guess_statistics = openpyxl.load_workbook("/Users/emvlp/Desktop/guess_the_number_statistics.xlsx")
        sheet_guess1 = excel_guess_statistics.active
        for row in sheet_guess1.iter_rows(min_row=2, max_col=2, values_only=True):
            print(f"Nombre del Jugador: {row[0]}, Resultado de la partida: {row[1]}")
    except FileNotFoundError:
        print("No hay estadísticas disponibles.")         

        
        
# Función para graficar estadísticas de los jugadores

def graficar_estadisticas():
    jugadores = list(estadisticas_jugadores.keys())
    partidas_ganadas = [estadisticas_jugadores[jugador]["Ganadas"] for jugador in jugadores]
    partidas_perdidas = [estadisticas_jugadores[jugador]["Perdidas"] for jugador in jugadores]

    # Bar plot - gráfico de barras 
    plt.bar(jugadores, partidas_ganadas, label="Partidas Ganadas", color="green")
    plt.bar(jugadores, partidas_perdidas, bottom=partidas_ganadas, label="Partidas Perdidas", color="red")

    plt.xlabel('Jugadores')
    plt.ylabel('Número de Partidas')
    plt.title('Estadísticas de Partidas')
    plt.legend()
    
    # Escala en números enteros
    if partidas_ganadas:
        plt.yticks(range(0, max(partidas_ganadas) + 1, 1))
        plt.ylim(0, max(partidas_ganadas) + 1)
    
    plt.show()
      
        

# Modo Solitario         

            
def modo_solitario():
    while True:
        print("Usted ha elegido la opción 1: Partida modo solitario.")
        print()
        dificultad = obtener_dificultad()

        if dificultad == 1:
            intentos = 20
        elif dificultad == 2:
            intentos = 12
        elif dificultad == 3:
            intentos = 5

        while True:
            try:
                secret_number = num_guess()
                if 1 <= secret_number <= 1000:
                    break
                else:
                    print("Número fuera del rango permitido (1 a 1000). Inténtalo nuevamente.")
            except ValueError:
                print("Entrada no válida. Ingrese un número.")

        print(f"Se ha elegido un número aleatorio. ¡Adivina el número! "
              f"Tienes {intentos} intentos.")

        while intentos > 0:
            try:
                guess = int(input("Adivina el número: "))
                if 1 <= guess <= 1000:
                    if guess == secret_number:
                        print("¡Felicidades! Has adivinado el número.")
                        break  # Sale del bucle de juego
                    elif guess < secret_number:
                        print("El número es mayor. Intenta nuevamente.")
                    else:
                        print("El número es menor. Intenta nuevamente.")
                    intentos -= 1
                    print(f"Tienes {intentos} intentos restantes. ¡Ánimo, vuelve a intentarlo! ")
                else:
                    print("Número fuera del rango permitido (1 a 1000). Inténtalo nuevamente.")   
            except ValueError:
                print("Entrada no válida. Ingrese un número.")
                
        nombre_jugador = input("Introduce tu nombre: ")
        print(f"¡Gracias por jugar a Guess the Number!")
        
        # Guardar estadísticas
        guardar_estadisticas(nombre_jugador, "Partida Ganada" if guess == secret_number else "Partida Perdida")

        volver_a_jugar = input("¿Quieres volver a jugar? (Sí/No): ").lower()
        if volver_a_jugar != 'si':
            break

# Modo dos Jugadores         

def modo_dos_jugadores():
    
    resultado_jugador1 = ""
    resultado_jugador2 = ""
    
    while True:
        print("Usted ha elegido la opción 2: Partida 2 jugadores.")
        print()
        dificultad = obtener_dificultad()

        if dificultad == 1:
            intentos = 20
        elif dificultad == 2:
            intentos = 12
        elif dificultad == 3:
            intentos = 5

        nombre_jugador1 = input("Jugador 1, introduce tu nombre: ")
            
        while True:
            try:
                secret_number = int(getpass.getpass("Jugador 1, ingresa el número secreto (entre 1 y 1000): "))
                if 1 <= secret_number <= 1000:
                    break  # Sale del bucle si el número es válido
                else:
                    print("Número fuera del rango permitido. Inténtalo nuevamente.")
            except ValueError:
                print("Entrada no válida. Ingrese un número.")

        print(f"Número secreto ingresado. ¡Que comience la adivinanza! Tienes {intentos} intentos.")

        while intentos > 0:
            try:
                guess = int(input("Jugador 2, adivina el número: "))
                if 1 <= guess <= 1000:
                    if guess == secret_number:
                        print("¡Felicidades! Jugador 2 ha adivinado el número. Jugador 1 ha perdido la partida.")
                        resultado_jugador1 = "Partida Perdida"
                        resultado_jugador2 = "Partida Ganada"
                        break  # Sale del bucle de juego
                    elif guess < secret_number:
                        print("El número es mayor. Intenta nuevamente.")
                    else:
                        print("El número es menor. Intenta nuevamente.")
                    intentos -= 1
                    print(f"Tienes {intentos} intentos restantes. ¡Ánimo, vuelve a intentarlo! ")
                else:
                    print("Número fuera del rango permitido (1 a 1000). Inténtalo nuevamente.")   
            except ValueError:
                print("Entrada no válida. Ingrese un número.")
                
        nombre_jugador2 = input("Introduce tu nombre: ")
        print(f"¡Gracias por jugar a Guess the Number! ")

        # Guardar estadísticas
        guardar_estadisticas(nombre_jugador1, resultado_jugador1)
        guardar_estadisticas(nombre_jugador2, resultado_jugador2)
        
        volver_a_jugar = input("¿Quieres volver a jugar? (Sí/No): ").lower()
        if volver_a_jugar != 'si':
            break


# Menú principal, Guess The Number 

while True:
    print("Bienvenido/a a Guess the Number's Game, escoga una de la opciones:")
    print()
    print("1. Partida modo solitario")
    print("2. Partida 2 jugadores")
    print("3. Estadística")
    print("4. Salir")

    menu = int(input("La opción a elegir es: "))
    print()
    if menu == 1:
        modo_solitario()
    elif menu == 2:
        modo_dos_jugadores()
    elif menu == 3:
        print("Usted ha elegido la opción 3: Estadística.")
        mostrar_estadisticas()
        graficar_estadisticas()  # función de graficar después de mostrar las estadísticas
    elif menu == 4:
        print("Usted ha elegido la opción 4: Salir. ¡Hasta pronto! ")
        break
    else:
        print("El valor", menu, "no se encuentra dentro de las cuatro opciones del menú. Inserte la opción nuevamente:")
        
